from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import shutil
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter




API_BASE = "https://clinicaltrials.gov/api/v2"


MASTER_TEMPLATE = "ClinicalTrials_Tracker_TEMPLATE_FINAL.xlsx"


OUTPUT_FILE = "ClinicalTrials_Tracker.xlsx"



DATA_HEADER_ROW = 1

OVERVIEW_SHEET_NAME = "Overview"

RESULTS_SHEET_SUFFIX = "_Results"


REQUIRED_BASE_SHEETS = [
    "README",
    "Dashboard",
    "Config",
    OVERVIEW_SHEET_NAME,
    "Change_Log",
]


CONFIG_REQUIRED_HEADERS = ["Asset", "Search Term", "NCT IDs", "Active"]


RESULTS_HEADER_FILL = PatternFill(
    fgColor="0F6B78",
    fill_type="solid",
)



OVERVIEW_COLUMNS = [
    "Asset",
    "NCT ID",
    "Official Study Title",
    "Brief Title",
    "Brief Summary",
    "Condition",
    "Intervention / Drug",
    "Lead Sponsor",
    "Collaborators",
    "Study Type",
    "Study Phase",
    "Start Date",
    "Primary Completion Date",
    "Study Completion Date",
    "Enrollment",
    "Recruitment Status",
    "Has Results",
    "Last Update Date",
    "ClinicalTrials.gov URL",
    "Results Snapshot Hash",
]

RESULT_COLUMNS = [
    "NCT ID",
    "Result Type",
    "Outcome / Measure",
    "Outcome Type",
    "Description",
    "Time Frame",
    "Group / Arm",
    "Participants",
    "Category",
    "Parameter Type",
    "Value",
    "Unit",
    "Spread",
    "Lower Limit",
    "Upper Limit",
    "P Value",
    "Statistical Test",
    "Event Type",
    "Organ System",
    "Event Term",
    "Number of Events",
    "Participants Affected",
    "Participants at Risk",
    "Milestone / Period",
    "Comment",
]

CHANGE_COLUMNS = [
    "Timestamp",
    "Asset",
    "NCT ID",
    "Change Type",
    "Field",
    "Old Value",
    "New Value",
]


TRACKED_FIELDS = [
    column
    for column in OVERVIEW_COLUMNS
    if column not in {
        "Asset",
        "ClinicalTrials.gov URL",
        "Results Snapshot Hash",
    }
]


logger = logging.getLogger("clinical_trials_tracker")




def setup_logging(log_path: Path) -> None:
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s"
    )

    file_handler = logging.FileHandler(
        log_path,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)




def clean_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, (dict, list)):
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
        )

    return str(value).strip()


def normalize_scalar(value: Any) -> Any:

    if value is None:
        return ""

    if isinstance(value, (int, float, bool)):
        return value

    text = str(value).strip()

    if not text:
        return ""

    compact = text.replace(",", "")

    try:
        if compact.lower() in {"nan", "inf", "-inf"}:
            return text

        if "." in compact or "e" in compact.lower():
            return float(compact)

        return int(compact)

    except ValueError:
        return text


def safe_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    return {}


def safe_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    return []


def join_unique(
    values: list[Any],
    separator: str = "\n",
) -> str:
    seen: set[str] = set()
    output: list[str] = []

    for value in values:
        text = clean_text(value)

        if text and text not in seen:
            seen.add(text)
            output.append(text)

    return separator.join(output)




def api_get(
    session: requests.Session,
    path: str,
    params: dict[str, Any] | None = None,
    retries: int = 3,
) -> dict[str, Any]:

    url = f"{API_BASE}{path}"
    last_error: Exception | None = None

    for attempt in range(1, retries + 1):

        try:
            response = session.get(
                url,
                params=params,
                timeout=60,
            )

            if response.status_code == 404:
                raise RuntimeError(
                    f"ClinicalTrials.gov returned 404: {url}"
                )

            response.raise_for_status()

            return response.json()

        except Exception as exc:

            last_error = exc

            if attempt < retries:
                delay = 2 ** (attempt - 1)

                logger.warning(
                    "Request failed: %s; retrying in %ss",
                    exc,
                    delay,
                )

                time.sleep(delay)

    raise RuntimeError(
        f"API request failed after {retries} attempts: {url}"
    ) from last_error


def get_data_timestamp(
    session: requests.Session,
) -> str:

    data = api_get(
        session,
        "/version",
    )

    return clean_text(
        data.get("dataTimestamp")
    )


def search_studies(
    session: requests.Session,
    search_term: str,
) -> list[str]:

    nct_ids: set[str] = set()
    page_token: str | None = None

    while True:

        params: dict[str, Any] = {
            "query.term": search_term,
            "pageSize": 1000,
        }

        if page_token:
            params["pageToken"] = page_token

        data = api_get(
            session,
            "/studies",
            params=params,
        )

        for study in safe_list(data.get("studies")):

            protocol = safe_dict(
                study.get("protocolSection")
            )

            identification = safe_dict(
                protocol.get("identificationModule")
            )

            nct_id = clean_text(
                identification.get("nctId")
            ).upper()

            if nct_id:
                nct_ids.add(nct_id)

        page_token = (
            clean_text(
                data.get("nextPageToken")
            )
            or None
        )

        if not page_token:
            break

    return sorted(nct_ids)


def get_study(
    session: requests.Session,
    nct_id: str,
) -> dict[str, Any]:

    return api_get(
        session,
        f"/studies/{nct_id}",
    )

def get_date(
    module: dict[str, Any],
    key: str,
) -> str:

    date_struct = safe_dict(
        module.get(key)
    )

    return clean_text(
        date_struct.get("date")
    )


def extract_interventions(
    module: dict[str, Any],
) -> str:

    values: list[str] = []

    for intervention in safe_list(
        module.get("interventions")
    ):

        intervention = safe_dict(
            intervention
        )

        name = clean_text(
            intervention.get("name")
        )

        other_names = [
            clean_text(value)
            for value in safe_list(
                intervention.get("otherNames")
            )
        ]

        combined = join_unique(
            [name] + other_names,
            separator="; ",
        )

        if combined:
            values.append(combined)

    return join_unique(
        values,
        separator="\n",
    )


def hash_results(
    results_section: dict[str, Any],
) -> str:

    payload = json.dumps(
        results_section,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )

    return hashlib.sha256(
        payload.encode("utf-8")
    ).hexdigest()


def extract_overview(
    study: dict[str, Any],
    asset: str,
) -> dict[str, Any]:

    protocol = safe_dict(
        study.get("protocolSection")
    )

    identification = safe_dict(
        protocol.get("identificationModule")
    )

    description = safe_dict(
        protocol.get("descriptionModule")
    )

    conditions = safe_dict(
        protocol.get("conditionsModule")
    )

    arms_interventions = safe_dict(
        protocol.get("armsInterventionsModule")
    )

    sponsor = safe_dict(
        protocol.get("sponsorCollaboratorsModule")
    )

    design = safe_dict(
        protocol.get("designModule")
    )

    status = safe_dict(
        protocol.get("statusModule")
    )

    lead_sponsor = safe_dict(
        sponsor.get("leadSponsor")
    )

    collaborators = [
        clean_text(
            safe_dict(item).get("name")
        )
        for item in safe_list(
            sponsor.get("collaborators")
        )
    ]

    enrollment_info = safe_dict(
        design.get("enrollmentInfo")
    )

    results_section = safe_dict(
        study.get("resultsSection")
    )


    has_results = (
        bool(study.get("hasResults"))
        or bool(results_section)
    )

    results_hash = (
        hash_results(results_section)
        if has_results
        else ""
    )

    nct_id = clean_text(
        identification.get("nctId")
    ).upper()

    return {
        "Asset": asset,
        "NCT ID": nct_id,
        "Official Study Title": clean_text(
            identification.get("officialTitle")
        ),
        "Brief Title": clean_text(
            identification.get("briefTitle")
        ),
        "Brief Summary": clean_text(
            description.get("briefSummary")
        ),
        "Condition": join_unique(
            safe_list(
                conditions.get("conditions")
            )
        ),
        "Intervention / Drug": extract_interventions(
            arms_interventions
        ),
        "Lead Sponsor": clean_text(
            lead_sponsor.get("name")
        ),
        "Collaborators": join_unique(
            collaborators
        ),
        "Study Type": clean_text(
            design.get("studyType")
        ),
        "Study Phase": join_unique(
            [
                clean_text(value)
                for value in safe_list(
                    design.get("phases")
                )
            ]
        ),
        "Start Date": get_date(
            status,
            "startDateStruct",
        ),
        "Primary Completion Date": get_date(
            status,
            "primaryCompletionDateStruct",
        ),
        "Study Completion Date": get_date(
            status,
            "completionDateStruct",
        ),
        "Enrollment": normalize_scalar(
            enrollment_info.get("count")
        ),
        "Recruitment Status": clean_text(
            status.get("overallStatus")
        ),
        "Has Results": (
            "TRUE"
            if has_results
            else "FALSE"
        ),
        "Last Update Date": get_date(
            status,
            "lastUpdatePostDateStruct",
        ),
        "ClinicalTrials.gov URL": (
            f"https://clinicaltrials.gov/study/{nct_id}"
            if nct_id
            else ""
        ),
        "Results Snapshot Hash": results_hash,
    }


def group_lookup(
    groups: Any,
) -> dict[str, str]:

    lookup: dict[str, str] = {}

    for group in safe_list(groups):

        group = safe_dict(group)

        group_id = clean_text(
            group.get("id")
        )

        title = clean_text(
            group.get("title")
        )

        if group_id:
            lookup[group_id] = title or group_id

    return lookup


def add_result(
    rows: list[dict[str, Any]],
    values: dict[str, Any],
) -> None:

    row = {
        column: ""
        for column in RESULT_COLUMNS
    }

    row.update(values)

    rows.append(row)




def extract_outcome_results(
    results_module: dict[str, Any],
    nct_id: str,
    rows: list[dict[str, Any]],
) -> None:

    for measure in safe_list(
        results_module.get("outcomeMeasures")
    ):

        measure = safe_dict(measure)

        measure_groups = group_lookup(
            measure.get("groups")
        )

        measure_title = clean_text(
            measure.get("title")
        )

        outcome_type = clean_text(
            measure.get("type")
        )

        description = clean_text(
            measure.get("description")
        )

        timeframe = clean_text(
            measure.get("timeFrame")
        )

        unit = clean_text(
            measure.get("unitOfMeasure")
        )

        # Outcome denominators.
        for denom in safe_list(
            measure.get("denoms")
        ):

            denom = safe_dict(denom)

            for count in safe_list(
                denom.get("counts")
            ):

                count = safe_dict(count)

                group_id = clean_text(
                    count.get("groupId")
                )

                add_result(
                    rows,
                    {
                        "NCT ID": nct_id,
                        "Result Type": "Outcome Denominator",
                        "Outcome / Measure": measure_title,
                        "Outcome Type": outcome_type,
                        "Description": description,
                        "Time Frame": timeframe,
                        "Group / Arm": measure_groups.get(
                            group_id,
                            group_id,
                        ),
                        "Participants": normalize_scalar(
                            count.get("value")
                        ),
                        "Unit": clean_text(
                            denom.get("units")
                        ),
                        "Comment": clean_text(
                            denom.get("unitsSelected")
                        ),
                    },
                )

        # Main measurement tables.
        for result_class in safe_list(
            measure.get("classes")
        ):

            result_class = safe_dict(
                result_class
            )

            class_title = clean_text(
                result_class.get("title")
            )

            for category in safe_list(
                result_class.get("categories")
            ):

                category = safe_dict(
                    category
                )

                category_title = (
                    clean_text(
                        category.get("title")
                    )
                    or class_title
                )

                for measurement in safe_list(
                    category.get("measurements")
                ):

                    measurement = safe_dict(
                        measurement
                    )

                    group_id = clean_text(
                        measurement.get("groupId")
                    )

                    if outcome_type == "PRIMARY":
                        result_type = "Primary"

                    elif outcome_type == "SECONDARY":
                        result_type = "Secondary"

                    else:
                        result_type = "Outcome"

                    add_result(
                        rows,
                        {
                            "NCT ID": nct_id,
                            "Result Type": result_type,
                            "Outcome / Measure": measure_title,
                            "Outcome Type": outcome_type,
                            "Description": description,
                            "Time Frame": timeframe,
                            "Group / Arm": measure_groups.get(
                                group_id,
                                group_id,
                            ),
                            "Category": category_title,
                            "Parameter Type": clean_text(
                                measure.get("paramType")
                            ),
                            "Value": normalize_scalar(
                                measurement.get("value")
                            ),
                            "Unit": unit,
                            "Spread": normalize_scalar(
                                measurement.get("spread")
                            ),
                            "Lower Limit": normalize_scalar(
                                measurement.get("lowerLimit")
                            ),
                            "Upper Limit": normalize_scalar(
                                measurement.get("upperLimit")
                            ),
                            "Comment": clean_text(
                                measurement.get("comment")
                            ),
                        },
                    )

       
        for analysis in safe_list(
            measure.get("analyses")
        ):

            analysis = safe_dict(
                analysis
            )

            comparison = (
                clean_text(
                    analysis.get("groupsDescription")
                )
                or clean_text(
                    analysis.get("comparisonGroupId")
                )
            )

            statistical_test = clean_text(
                analysis.get("statisticalMethod")
            )

            comment_parts = [
                text
                for text in [
                    clean_text(
                        analysis.get(
                            "otherAnalysisDescription"
                        )
                    ),
                    comparison,
                ]
                if text
            ]

            add_result(
                rows,
                {
                    "NCT ID": nct_id,
                    "Result Type": "Statistical Analysis",
                    "Outcome / Measure": measure_title,
                    "Outcome Type": outcome_type,
                    "Description": description,
                    "Time Frame": timeframe,
                    "Group / Arm": comparison,
                    "Parameter Type": clean_text(
                        analysis.get("paramType")
                    ),
                    "Value": normalize_scalar(
                        analysis.get("paramValue")
                    ),
                    "P Value": normalize_scalar(
                        analysis.get("pValue")
                    ),
                    "Statistical Test": statistical_test,
                    "Comment": " | ".join(
                        comment_parts
                    ),
                },
            )



def extract_baseline_results(
    baseline_module: dict[str, Any],
    nct_id: str,
    rows: list[dict[str, Any]],
) -> None:

    groups = group_lookup(
        baseline_module.get("groups")
    )

    for measure in safe_list(
        baseline_module.get("measures")
    ):

        measure = safe_dict(measure)

        title = clean_text(
            measure.get("title")
        )

        description = (
            clean_text(
                measure.get("description")
            )
            or clean_text(
                measure.get(
                    "populationDescription"
                )
            )
        )

        unit = clean_text(
            measure.get("unitOfMeasure")
        )

        parameter_type = clean_text(
            measure.get("paramType")
        )

        # Baseline denominators.
        for denom in safe_list(
            measure.get("denoms")
        ):

            denom = safe_dict(denom)

            for count in safe_list(
                denom.get("counts")
            ):

                count = safe_dict(count)

                group_id = clean_text(
                    count.get("groupId")
                )

                add_result(
                    rows,
                    {
                        "NCT ID": nct_id,
                        "Result Type": "Baseline Denominator",
                        "Outcome / Measure": title,
                        "Description": description,
                        "Group / Arm": groups.get(
                            group_id,
                            group_id,
                        ),
                        "Participants": normalize_scalar(
                            count.get("value")
                        ),
                        "Unit": clean_text(
                            denom.get("units")
                        ),
                        "Category": "Denominator",
                    },
                )

        # Baseline measurement tables.
        for result_class in safe_list(
            measure.get("classes")
        ):

            result_class = safe_dict(
                result_class
            )

            class_title = clean_text(
                result_class.get("title")
            )

            for category in safe_list(
                result_class.get("categories")
            ):

                category = safe_dict(
                    category
                )

                category_title = (
                    clean_text(
                        category.get("title")
                    )
                    or class_title
                )

                for measurement in safe_list(
                    category.get("measurements")
                ):

                    measurement = safe_dict(
                        measurement
                    )

                    group_id = clean_text(
                        measurement.get("groupId")
                    )

                    add_result(
                        rows,
                        {
                            "NCT ID": nct_id,
                            "Result Type": "Baseline",
                            "Outcome / Measure": title,
                            "Description": description,
                            "Group / Arm": groups.get(
                                group_id,
                                group_id,
                            ),
                            "Category": category_title,
                            "Parameter Type": parameter_type,
                            "Value": normalize_scalar(
                                measurement.get("value")
                            ),
                            "Unit": unit,
                            "Spread": normalize_scalar(
                                measurement.get("spread")
                            ),
                            "Lower Limit": normalize_scalar(
                                measurement.get("lowerLimit")
                            ),
                            "Upper Limit": normalize_scalar(
                                measurement.get("upperLimit")
                            ),
                            "Comment": clean_text(
                                measurement.get("comment")
                            ),
                        },
                    )



def extract_participant_flow(
    flow_module: dict[str, Any],
    nct_id: str,
    rows: list[dict[str, Any]],
) -> None:

    groups = group_lookup(
        flow_module.get("groups")
    )

    recruitment_details = clean_text(
        flow_module.get(
            "recruitmentDetails"
        )
    )

    pre_assignment_details = clean_text(
        flow_module.get(
            "preAssignmentDetails"
        )
    )

    if recruitment_details:

        add_result(
            rows,
            {
                "NCT ID": nct_id,
                "Result Type": "Participant Flow",
                "Description": recruitment_details,
                "Comment": "Recruitment Details",
            },
        )

    if pre_assignment_details:

        add_result(
            rows,
            {
                "NCT ID": nct_id,
                "Result Type": "Participant Flow",
                "Description": pre_assignment_details,
                "Comment": "Pre-assignment Details",
            },
        )

    for period in safe_list(
        flow_module.get("periods")
    ):

        period = safe_dict(period)

        period_title = clean_text(
            period.get("title")
        )

        for milestone in safe_list(
            period.get("milestones")
        ):

            milestone = safe_dict(
                milestone
            )

            milestone_type = clean_text(
                milestone.get("type")
            )

            for achievement in safe_list(
                milestone.get("achievements")
            ):

                achievement = safe_dict(
                    achievement
                )

                group_id = clean_text(
                    achievement.get("groupId")
                )

                add_result(
                    rows,
                    {
                        "NCT ID": nct_id,
                        "Result Type": "Participant Flow",
                        "Group / Arm": groups.get(
                            group_id,
                            group_id,
                        ),
                        "Participants": normalize_scalar(
                            achievement.get(
                                "numSubjects"
                            )
                        ),
                        "Milestone / Period": (
                            f"{period_title} | "
                            f"{milestone_type}"
                        ).strip(" |"),
                        "Comment": clean_text(
                            achievement.get("comment")
                        ),
                    },
                )

        for drop_withdraw in safe_list(
            period.get("dropWithdraws")
        ):

            drop_withdraw = safe_dict(
                drop_withdraw
            )

            drop_type = clean_text(
                drop_withdraw.get("type")
            )

            for reason in safe_list(
                drop_withdraw.get("reasons")
            ):

                reason = safe_dict(
                    reason
                )

                group_id = clean_text(
                    reason.get("groupId")
                )

                add_result(
                    rows,
                    {
                        "NCT ID": nct_id,
                        "Result Type": (
                            "Participant Flow - Withdrawal"
                        ),
                        "Group / Arm": groups.get(
                            group_id,
                            group_id,
                        ),
                        "Participants": normalize_scalar(
                            reason.get("numSubjects")
                        ),
                        "Event Type": drop_type,
                        "Comment": clean_text(
                            reason.get("reason")
                        ),
                        "Milestone / Period": period_title,
                    },
                )




def extract_adverse_events(
    adverse_module: dict[str, Any],
    nct_id: str,
    rows: list[dict[str, Any]],
) -> None:

    event_groups = {}

    for group in safe_list(
        adverse_module.get("eventGroups")
    ):

        group = safe_dict(
            group
        )

        group_id = clean_text(
            group.get("id")
        )

        if group_id:
            event_groups[group_id] = clean_text(
                group.get("title")
            ) or group_id

    timeframe = clean_text(
        adverse_module.get("timeFrame")
    )


    for group in safe_list(
        adverse_module.get("eventGroups")
    ):

        group = safe_dict(
            group
        )

        group_id = clean_text(
            group.get("id")
        )

        group_title = clean_text(
            group.get("title")
        )

        summary_types = [
            (
                "Other",
                "otherNumAtRisk",
                "otherNumAffected",
            ),
            (
                "Serious",
                "seriousNumAtRisk",
                "seriousNumAffected",
            ),
            (
                "Deaths",
                "deathsNumAtRisk",
                "deathsNumAffected",
            ),
        ]

        for event_type, at_risk_key, affected_key in summary_types:

            if (
                at_risk_key in group
                or affected_key in group
            ):

                add_result(
                    rows,
                    {
                        "NCT ID": nct_id,
                        "Result Type": "Adverse Event Summary",
                        "Time Frame": timeframe,
                        "Group / Arm": (
                            group_title
                            or group_id
                        ),
                        "Event Type": event_type,
                        "Participants Affected": normalize_scalar(
                            group.get(
                                affected_key
                            )
                        ),
                        "Participants at Risk": normalize_scalar(
                            group.get(
                                at_risk_key
                            )
                        ),
                        "Comment": clean_text(
                            group.get(
                                "description"
                            )
                        ),
                    },
                )

    # Detailed adverse events.
    for event_key, result_type in [
        (
            "seriousEvents",
            "Adverse Event - Serious",
        ),
        (
            "otherEvents",
            "Adverse Event - Other",
        ),
    ]:

        for event in safe_list(
            adverse_module.get(event_key)
        ):

            event = safe_dict(
                event
            )

            event_term = clean_text(
                event.get("term")
            )

            organ_system = clean_text(
                event.get("organSystem")
            )

            assessment_type = clean_text(
                event.get("assessmentType")
            )

            notes = clean_text(
                event.get("notes")
            )

            for stat in safe_list(
                event.get("stats")
            ):

                stat = safe_dict(
                    stat
                )

                group_id = clean_text(
                    stat.get("groupId")
                )

                add_result(
                    rows,
                    {
                        "NCT ID": nct_id,
                        "Result Type": result_type,
                        "Time Frame": timeframe,
                        "Group / Arm": event_groups.get(
                            group_id,
                            group_id,
                        ),
                        "Event Type": (
                            assessment_type
                            or event_key
                        ),
                        "Organ System": organ_system,
                        "Event Term": event_term,
                        "Number of Events": normalize_scalar(
                            stat.get(
                                "numEvents"
                            )
                        ),
                        "Participants Affected": normalize_scalar(
                            stat.get(
                                "numAffected"
                            )
                        ),
                        "Participants at Risk": normalize_scalar(
                            stat.get(
                                "numAtRisk"
                            )
                        ),
                        "Comment": notes,
                    },
                )



def extract_results(
    study: dict[str, Any],
    nct_id: str,
) -> list[dict[str, Any]]:

    results_section = safe_dict(
        study.get("resultsSection")
    )

    rows: list[dict[str, Any]] = []

    if not results_section:
        return rows

    outcome_module = safe_dict(
        results_section.get(
            "outcomeMeasuresModule"
        )
    )

    baseline_module = safe_dict(
        results_section.get(
            "baselineCharacteristicsModule"
        )
    )

    flow_module = safe_dict(
        results_section.get(
            "participantFlowModule"
        )
    )

    adverse_module = safe_dict(
        results_section.get(
            "adverseEventsModule"
        )
    )

    extract_outcome_results(
        outcome_module,
        nct_id,
        rows,
    )

    extract_baseline_results(
        baseline_module,
        nct_id,
        rows,
    )

    extract_participant_flow(
        flow_module,
        nct_id,
        rows,
    )

    extract_adverse_events(
        adverse_module,
        nct_id,
        rows,
    )

    return rows



def sanitize_sheet_name(name: str) -> str:
    """
    Make an arbitrary asset name safe to use as an Excel sheet name:
    strip characters Excel forbids in sheet names and cap the length.
    """

    cleaned = re.sub(
        r'[:\\/?*\[\]]',
        "_",
        name,
    ).strip()

    return cleaned[:31] if cleaned else "Sheet"


def results_sheet_name(asset: str) -> str:

    max_asset_length = 31 - len(RESULTS_SHEET_SUFFIX)

    base = sanitize_sheet_name(asset)[:max_asset_length]

    return f"{base}{RESULTS_SHEET_SUFFIX}"


def find_config_header_row(
    ws,
    max_scan: int = 20,
) -> int:
 

    for row in ws.iter_rows(
        min_row=1,
        max_row=max_scan,
    ):

        values = {
            clean_text(cell.value)
            for cell in row
        }

        if {"Asset", "Active"}.issubset(values):
            return row[0].row

    raise ValueError(
        "Could not locate the Config header row "
        "(expected a row containing 'Asset' and 'Active')."
    )


def parse_config(
    ws,
) -> list[tuple[str, str, str]]:

    header_row = find_config_header_row(ws)

    headers = {
        clean_text(cell.value): cell.column
        for cell in ws[header_row]
    }

    required = {
        "Asset",
        "Active",
    }

    missing = required - set(headers)

    if missing:
        raise ValueError(
            "Config is missing columns: "
            + ", ".join(sorted(missing))
        )

    items: list[
        tuple[str, str, str]
    ] = []

    for row in ws.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):

        asset = clean_text(
            row[headers["Asset"] - 1]
        )

        search_term = (
            clean_text(
                row[
                    headers["Search Term"] - 1
                ]
            )
            if "Search Term" in headers
            else ""
        )

        direct_ncts = (
            clean_text(
                row[
                    headers["NCT IDs"] - 1
                ]
            )
            if "NCT IDs" in headers
            else ""
        )

        active = row[
            headers["Active"] - 1
        ]

        if not asset:
            continue

        if str(active).strip().lower() in {
            "false",
            "0",
            "no",
            "n",
            "inactive",
        }:
            continue

        if not search_term and not direct_ncts:
            continue

        items.append(
            (
                asset,
                search_term,
                direct_ncts,
            )
        )

    return items




def ensure_template_structure(
    template_path: Path,
) -> None:

    if not template_path.exists():
        raise FileNotFoundError(
            f"Master template not found: "
            f"{template_path}"
        )

    wb = load_workbook(
        template_path,
        read_only=True,
        data_only=False,
    )

    missing_sheets = [
        name
        for name in REQUIRED_BASE_SHEETS
        if name not in wb.sheetnames
    ]

    if missing_sheets:
        raise ValueError(
            "Template is missing required worksheets: "
            + ", ".join(missing_sheets)
        )

    config_header_row = find_config_header_row(
        wb["Config"]
    )

    config_headers = [
        clean_text(cell.value)
        for cell in wb["Config"][
            config_header_row
        ]
    ]

    if config_headers[:4] != CONFIG_REQUIRED_HEADERS:
        raise ValueError(
            "Config header row does not match the expected template "
            f"(found {config_headers[:4]})."
        )

    overview_headers = [
        clean_text(cell.value)
        for cell in wb[OVERVIEW_SHEET_NAME][
            DATA_HEADER_ROW
        ]
    ]

    if overview_headers[: len(OVERVIEW_COLUMNS)] != OVERVIEW_COLUMNS:
        raise ValueError(
            f"{OVERVIEW_SHEET_NAME} headers do not match the template."
        )

    change_headers = [
        clean_text(cell.value)
        for cell in wb["Change_Log"][
            DATA_HEADER_ROW
        ]
    ]

    if change_headers[: len(CHANGE_COLUMNS)] != CHANGE_COLUMNS:
        raise ValueError(
            "Change_Log headers do not match the template."
        )


def workbook_matches_template(
    workbook_path: Path,
) -> bool:

    try:

        wb = load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
        )

        missing_sheets = [
            name
            for name in REQUIRED_BASE_SHEETS
            if name not in wb.sheetnames
        ]

        if missing_sheets:
            wb.close()
            return False

        config_header_row = find_config_header_row(
            wb["Config"]
        )

        config_headers = [
            clean_text(cell.value)
            for cell in wb["Config"][
                config_header_row
            ]
        ]

        if config_headers[:4] != CONFIG_REQUIRED_HEADERS:
            wb.close()
            return False

        overview_headers = [
            clean_text(cell.value)
            for cell in wb[OVERVIEW_SHEET_NAME][1]
        ]

        if overview_headers[: len(OVERVIEW_COLUMNS)] != OVERVIEW_COLUMNS:
            wb.close()
            return False

        change_headers = [
            clean_text(cell.value)
            for cell in wb["Change_Log"][1]
        ]

        if change_headers[: len(CHANGE_COLUMNS)] != CHANGE_COLUMNS:
            wb.close()
            return False

        # Any per-asset Results sheets that already exist (from a
        # prior run) must still match the current schema. Sheets for
        # assets that haven't been processed yet simply won't exist,
        # and that's fine - they're created on demand.
        for name in wb.sheetnames:

            if (
                name.endswith(RESULTS_SHEET_SUFFIX)
                and name not in REQUIRED_BASE_SHEETS
            ):

                results_headers = [
                    clean_text(cell.value)
                    for cell in wb[name][1]
                ]

                if results_headers[: len(RESULT_COLUMNS)] != RESULT_COLUMNS:
                    wb.close()
                    return False

        wb.close()
        return True

    except Exception:
        return False


def make_output_from_template(
    template_path: Path,
    output_path: Path,
) -> None:

    if (
        output_path.exists()
        and workbook_matches_template(
            output_path
        )
    ):
        return

    if output_path.exists():

        backup_path = output_path.with_name(
            output_path.stem
            + f"_backup_{datetime.now():%Y%m%d_%H%M%S}"
            + output_path.suffix
        )

        shutil.move(
            output_path,
            backup_path,
        )

        logger.warning(
            "Existing output did not match "
            "the current template. "
            "Backed it up to %s",
            backup_path,
        )

    shutil.copy2(
        template_path,
        output_path,
    )




def row_records(
    ws,
    columns: list[str],
    header_row: int = 1,
) -> list[dict[str, Any]]:

    headers = {
        clean_text(cell.value): cell.column
        for cell in ws[header_row]
    }

    records: list[
        dict[str, Any]
    ] = []

    for row in ws.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):

        if not any(
            value not in (None, "")
            for value in row
        ):
            continue

        record = {}

        for column in columns:

            column_index = headers.get(
                column
            )

            if column_index:
                record[column] = row[
                    column_index - 1
                ]
            else:
                record[column] = ""

        records.append(record)

    return records


def index_overview(
    ws,
) -> dict[str, dict[str, Any]]:

    records = row_records(
        ws,
        OVERVIEW_COLUMNS,
    )

    indexed: dict[
        str,
        dict[str, Any]
    ] = {}

    for record in records:

        nct_id = clean_text(
            record.get("NCT ID")
        ).upper()

        if nct_id:
            indexed[nct_id] = record

    return indexed


def index_results(
    ws,
) -> list[dict[str, Any]]:

    return row_records(
        ws,
        RESULT_COLUMNS,
    )


def write_table(
    ws,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:


    if ws.max_row > 1:
        ws.delete_rows(
            2,
            ws.max_row - 1,
        )

    for row_data in rows:

        ws.append(
            [
                row_data.get(
                    column,
                    "",
                )
                for column in headers
            ]
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(headers))}"
        f"{max(1, len(rows) + 1)}"
    )

    ws.sheet_view.showGridLines = False

    # Imported data is shown in green.
    for row in ws.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

            cell.font = Font(
                color="008000"
            )

    # Keep header readable.
    for cell in ws[1]:

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )




def get_overview_sheet(
    wb,
):
    return wb[OVERVIEW_SHEET_NAME]


def get_or_create_results_sheet(
    wb,
    asset: str,
):
  

    sheet_name = results_sheet_name(asset)

    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    ws.append(RESULT_COLUMNS)

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFFFF",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.fill = RESULTS_HEADER_FILL

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

 
    if "Change_Log" in wb.sheetnames:

        target_index = wb.sheetnames.index(
            "Change_Log"
        )

        wb._sheets.remove(ws)
        wb._sheets.insert(
            target_index,
            ws,
        )

    logger.info(
        "Created new Results worksheet: %s",
        sheet_name,
    )

    return ws


def append_change(
    ws,
    asset: str,
    nct_id: str,
    change_type: str,
    field: str,
    old_value: Any,
    new_value: Any,
) -> None:

    ws.append(
        [
            datetime.now(
                timezone.utc
            ).astimezone().isoformat(
                timespec="seconds"
            ),
            asset,
            nct_id,
            change_type,
            field,
            clean_text(old_value),
            clean_text(new_value),
        ]
    )


def add_changes(
    change_ws,
    asset: str,
    nct_id: str,
    old_record: dict[str, Any] | None,
    new_record: dict[str, Any],
    results_changed: bool,
) -> None:

    if old_record is None:

        append_change(
            change_ws,
            asset,
            nct_id,
            "NEW STUDY",
            "",
            "",
            "Study discovered",
        )

        return

    # Compare protocol/overview fields.
    for field in TRACKED_FIELDS:

        old_value = clean_text(
            old_record.get(field)
        )

        new_value = clean_text(
            new_record.get(field)
        )

        if old_value != new_value:

            append_change(
                change_ws,
                asset,
                nct_id,
                "UPDATED",
                field,
                old_value,
                new_value,
            )


    if results_changed:

        append_change(
            change_ws,
            asset,
            nct_id,
            "RESULTS UPDATED",
            "Results Snapshot Hash",
            old_record.get(
                "Results Snapshot Hash",
                "",
            ),
            new_record.get(
                "Results Snapshot Hash",
                "",
            ),
        )




def update_asset_sheet(
    wb,
    asset: str,
    successful_studies: dict[
        str,
        dict[str, Any],
    ],
    change_log_ws,
) -> int:

    overview_ws = get_overview_sheet(wb)

    results_ws = get_or_create_results_sheet(
        wb,
        asset,
    )

    old_overview = index_overview(
        overview_ws
    )

    old_results = index_results(
        results_ws
    )

    merged_overview = dict(
        old_overview
    )

    merged_results = list(
        old_results
    )

    for nct_id, study_data in (
        successful_studies.items()
    ):

        overview = study_data[
            "overview"
        ]

        new_results = study_data[
            "results"
        ]

        old = old_overview.get(
            nct_id
        )

        old_hash = (
            clean_text(
                old.get(
                    "Results Snapshot Hash"
                )
            )
            if old
            else ""
        )

        new_hash = clean_text(
            overview.get(
                "Results Snapshot Hash"
            )
        )


        try:

            add_changes(
                change_log_ws,
                asset,
                nct_id,
                old,
                overview,
                old_hash != new_hash,
            )

        except Exception as exc:

            logger.error(
                "Change Log error for %s: %s",
                nct_id,
                exc,
            )

        merged_overview[
            nct_id
        ] = overview

  
        merged_results = [
            row
            for row in merged_results
            if clean_text(
                row.get("NCT ID")
            ).upper()
            != nct_id
        ]

        merged_results.extend(
            new_results
        )

    overview_rows = [
        merged_overview[nct]
        for nct in sorted(
            merged_overview
        )
    ]

    write_table(
        overview_ws,
        OVERVIEW_COLUMNS,
        overview_rows,
    )

    write_table(
        results_ws,
        RESULT_COLUMNS,
        merged_results,
    )

    return len(
        merged_overview
    )



def update_dashboard(
    wb,
    config_items: list[
        tuple[str, str, str]
    ],
    data_timestamp: str = "",
) -> None:

    ws = wb["Dashboard"]

    overview_ws = get_overview_sheet(wb)

    unique_ncts: set[str] = set()
    studies_with_results = 0
    result_rows = 0


## JX: change here, all studies share the same worksheet
## modify here if want to seperate later
    for record in row_records(
        overview_ws,
        OVERVIEW_COLUMNS,
    ):

        nct_id = clean_text(
            record.get("NCT ID")
        ).upper()

        if nct_id:
            unique_ncts.add(
                nct_id
            )

        if (
            clean_text(
                record.get("Has Results")
            ).upper()
            == "TRUE"
        ):
            studies_with_results += 1

    for asset, _, _ in config_items:

        sheet_name = results_sheet_name(asset)

        if sheet_name not in wb.sheetnames:
            continue

        result_rows += max(
            0,
            wb[sheet_name].max_row - 1,
        )

    metrics = {
        "Active Assets": len(
            config_items
        ),
        "Unique Studies": len(
            unique_ncts
        ),
        "Studies With Posted Results": (
            studies_with_results
        ),
        "Total Result Rows": result_rows,
        "Last API Data Timestamp": (
            data_timestamp
            or datetime.now().astimezone().isoformat(
                timespec="seconds"
            )
        ),
    }

    for row_number in range(4, 9):

        label = clean_text(
            ws.cell(
                row=row_number,
                column=1,
            ).value
        )

        if label in metrics:

            ws.cell(
                row=row_number,
                column=2,
            ).value = metrics[label]

            ws.cell(
                row=row_number,
                column=2,
            ).font = Font(
                bold=True,
                color="008080",
            )




def main() -> int:

    parser = argparse.ArgumentParser(
        description=(
            "ClinicalTrials.gov asset tracker"
        )
    )

    parser.add_argument(
        "--template",
        default=MASTER_TEMPLATE,
    )

    parser.add_argument(
        "--output",
        default=OUTPUT_FILE,
    )

    parser.add_argument(
        "--log",
        default="clinical_trials_tracker.log",
    )

    args = parser.parse_args()

    base_dir = Path(
        __file__
    ).resolve().parent

    template_path = Path(
        args.template
    )

    output_path = Path(
        args.output
    )

    log_path = Path(
        args.log
    )

    if not template_path.is_absolute():
        template_path = (
            base_dir
            / template_path
        )

    if not output_path.is_absolute():
        output_path = (
            base_dir
            / output_path
        )

    if not log_path.is_absolute():
        log_path = (
            base_dir
            / log_path
        )

    setup_logging(
        log_path
    )


    ensure_template_structure(
        template_path
    )

    make_output_from_template(
        template_path,
        output_path,
    )

    wb = load_workbook(
        output_path
    )

    config_items = parse_config(
        wb["Config"]
    )

    change_log_ws = wb[
        "Change_Log"
    ]

    session = requests.Session()

    session.headers.update(
        {
            "User-Agent": (
                "ClinicalTrialsTracker/3.0"
            )
        }
    )

    data_timestamp = ""

    try:

        data_timestamp = (
            get_data_timestamp(
                session
            )
        )

        logger.info(
            "ClinicalTrials.gov "
            "dataTimestamp=%s",
            data_timestamp,
        )

    except Exception as exc:

        logger.warning(
            "Could not read ClinicalTrials.gov "
            "dataTimestamp: %s",
            exc,
        )

    total_studies = 0

    for (
        asset,
        search_term,
        direct_ncts_text,
    ) in config_items:

        logger.info(
            "Searching asset=%s using term=%s",
            asset,
            search_term
            or "<direct NCT IDs>",
        )

  
        discovered_ids = (
            search_studies(
                session,
                search_term,
            )
            if search_term
            else []
        )

   
        direct_ids = [
            item.strip().upper()
            for item in (
                direct_ncts_text
                .replace(";", ",")
                .split(",")
            )
            if item.strip()
        ]

        nct_ids = sorted(
            set(discovered_ids)
            | set(direct_ids)
        )

        logger.info(
            "Found %d unique studies for %s",
            len(nct_ids),
            asset,
        )

        successful: dict[
            str,
            dict[str, Any],
        ] = {}

        for index, nct_id in enumerate(
            nct_ids,
            start=1,
        ):

            try:

                
                study = get_study(
                    session,
                    nct_id,
                )

                overview = extract_overview(
                    study,
                    asset,
                )

                results = extract_results(
                    study,
                    nct_id,
                )

                results_section_exists = bool(
                    safe_dict(
                        study.get(
                            "resultsSection"
                        )
                    )
                )

             
                logger.info(
                    (
                        "%s: "
                        "hasResults=%s; "
                        "resultsSection=%s; "
                        "resultRows=%d"
                    ),
                    nct_id,
                    overview["Has Results"],
                    results_section_exists,
                    len(results),
                )

            
                if (
                    overview[
                        "Has Results"
                    ]
                    == "TRUE"
                    and not results
                ):

                    logger.warning(
                        (
                            "%s has Results posted "
                            "but parser extracted "
                            "0 rows"
                        ),
                        nct_id,
                    )

                successful[
                    nct_id
                ] = {
                    "overview": overview,
                    "results": results,
                }

            except Exception as exc:

             
                logger.error(
                    "Failed to retrieve %s: %s",
                    nct_id,
                    exc,
                )

            if index % 25 == 0:

                logger.info(
                    "Processed %d/%d studies for %s",
                    index,
                    len(nct_ids),
                    asset,
                )

        update_asset_sheet(
            wb,
            asset,
            successful,
            change_log_ws,
        )

    total_studies = len(
        index_overview(
            get_overview_sheet(wb)
        )
    )

    update_dashboard(
        wb,
        config_items,
        data_timestamp,
    )

    wb.save(
        output_path
    )

    logger.info(
        "Saved workbook: %s",
        output_path,
    )

    logger.info(
        "Tracked %d study records across %d assets",
        total_studies,
        len(config_items),
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )